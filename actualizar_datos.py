"""
actualizar_datos.py
===================================================================
Este es el motor de mi dashboard. Corre solo todos los días por GitHub
Actions, jala todo de las APIs y deja listo el datos.json que lee el
index.html. Lo que estoy bajando:

  - Henry Hub (gas natural)        -> Yahoo NG=F, y si falla, EIA spot
  - Inflación EE. UU. (CPI + YoY)  -> FRED
  - S&P 500                        -> FRED
  - WTI                            -> FRED
  - Brent                          -> FRED
  - 13 divisas vs. USD             -> Frankfurter (BCE)
  - Acereras (acciones base 100)   -> Twelve Data + Yahoo

Todo arranca en 2021-01-01 para tener una base común.

Las divisas las saco de Frankfurter porque con base=USD ya vienen como
"unidades por 1 USD" (18 pesos, 157 yenes...), sin tener que invertir nada,
y salen el mismo día hábil ~16:00 hora de Europa. Más fresco que FRED.

Cada corrida me deja en la carpeta `resultados`:
  - mercados_AAAA-MM-DD.xlsx  y  mercados_reciente.xlsx
  - graficos_AAAA-MM-DD.png   y  graficos_reciente.png
  - registro.log

-------------------------------------------------------------------
Si lo corro local, una sola vez:
    pip install requests pandas matplotlib openpyxl yfinance

Las llaves (EIA, FRED, Twelve Data):
  - En mi PC: un .env aquí mismo con
        EIA_API_KEY=...
        FRED_API_KEY=...
        TWELVEDATA_API_KEY=...
    (para eso también va `pip install python-dotenv`)
  - En GitHub Actions van como Secrets del repo y el workflow las mete
    como variables de entorno. Ahí no necesito ni .env ni dotenv.
  Frankfurter y Yahoo no piden llave. Si un día se me cae la de Twelve Data
  no pasa nada: las acereras de China/India (Yahoo) siguen entrando y solo
  se me caen las que dependen de Twelve Data (los ADRs gringos).
===================================================================
"""

import os
import sys
import json
import time
import datetime as dt
from pathlib import Path

import requests
import pandas as pd
import matplotlib
matplotlib.use("Agg")                 # modo sin ventana, solo me guarda el PNG
import matplotlib.pyplot as plt

# dotenv es opcional: en mi PC me lee el .env, y en GitHub Actions las llaves
# ya vienen de los Secrets, así que si no está instalado me lo salto y ya.
try:
    from dotenv import load_dotenv
    load_dotenv()
except ModuleNotFoundError:
    pass

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

# ------------------------------------------------------------------
# 0) Config
# ------------------------------------------------------------------
EIA_API_KEY = os.getenv("EIA_API_KEY")
FRED_API_KEY = os.getenv("FRED_API_KEY")
# Las acciones de acereras las jalo de Twelve Data. El workflow ya me mete
# este Secret como variable de entorno; en mi PC va en el .env como las demás.
TWELVEDATA_API_KEY = os.getenv("TWELVEDATA_API_KEY")

FECHA_INICIO = "2021-01-01"           # de aquí arranco todo, para tener base pareja

CARPETA = Path(__file__).resolve().parent / "resultados"
CARPETA.mkdir(exist_ok=True)
HOY = dt.date.today().isoformat()
LOG = CARPETA / "registro.log"

# Henry Hub en $/MMBtu (viene de EIA)
COL_HH = "Henry Hub ($/MMBtu)"
FMT_HH = "#,##0.00"

# Series diarias de FRED: id -> (nombre de columna, formato)
FRED_DIARIAS = {
    "SP500":        ("S&P 500",       "#,##0.00"),
    "DCOILWTICO":   ("WTI ($/bbl)",   "#,##0.00"),
    "DCOILBRENTEU": ("Brent ($/bbl)", "#,##0.00"),
}

# Divisas por Frankfurter (BCE). base=USD -> "unidades por 1 USD".
# código ISO -> (nombre de columna, formato)
MONEDAS = {
    "EUR": ("EUR por USD", "#,##0.0000"),
    "JPY": ("JPY por USD", "#,##0.00"),
    "CNY": ("CNY por USD", "#,##0.0000"),
    "GBP": ("GBP por USD", "#,##0.0000"),
    "MXN": ("MXN por USD", "#,##0.0000"),
    "CAD": ("CAD por USD", "#,##0.0000"),
    "BRL": ("BRL por USD", "#,##0.0000"),
    "AUD": ("AUD por USD", "#,##0.0000"),
    "INR": ("INR por USD", "#,##0.00"),
    "KRW": ("KRW por USD", "#,##0.00"),
    "COP": ("COP por USD", "#,##0.00"),
    "HKD": ("HKD por USD", "#,##0.00"),
    "TRY": ("TRY por USD", "#,##0.00"),
}

# Saneamiento: rangos realistas por serie. Si un dato cae fuera del rango lo
# trato como basura de la fuente (picos falsos) y lo dejo vacío; la línea de la
# gráfica simplemente lo brinca. Ojo: el techo de Henry Hub (25) me deja vivo el
# pico real del invierno 2021 (tormenta Uri) pero tira valores imposibles tipo
# ~30. Si algún día me quedan cortos estos números, los aflojo aquí.
RANGOS_VALIDOS = {
    COL_HH:           (0.5, 25.0),    # Henry Hub $/MMBtu
    "WTI ($/bbl)":    (5.0, 250.0),   # WTI $/bbl
    "Brent ($/bbl)":  (5.0, 250.0),   # Brent $/bbl
}


def sanear(df, col):
    """Deja en vacío los valores de 'col' que se salen de su rango realista."""
    if col not in RANGOS_VALIDOS or col not in df.columns:
        return df
    lo, hi = RANGOS_VALIDOS[col]
    mask = df[col].notna() & ((df[col] < lo) | (df[col] > hi))
    n = int(mask.sum())
    if n:
        log(f"  Saneando {col}: {n} valor(es) fuera de [{lo}, {hi}] -> vacío")
        df.loc[mask, col] = float("nan")
    return df


# Acereras que cotizan en bolsa, vía Twelve Data (símbolo -> nombre).
# Con el plan GRATIS de Twelve Data solo me entran bien las acciones listadas en
# EE. UU. (NYSE/Nasdaq), así que para las de afuera uso su ADR gringo. Las que
# solo cotizan fuera (China e India) las jalo por Yahoo (ver ACERERAS_YAHOO más
# abajo). Si un símbolo no contesta, lo brinco y lo apunto en "fallaron" — el
# índice equal-weight aguanta que entren y salgan empresas sin despeinarse.
ACERERAS_MUNDIAL = {
    "MT":     "ArcelorMittal",     # NYSE
    "PKX":    "POSCO",             # NYSE (ADR)
    "TX":     "Ternium",           # NYSE
    "GGB":    "Gerdau",            # NYSE
    "NPSCY":  "Nippon Steel",      # OTC (ADR) — a veces pide plan de pago
    "TKAMY":  "thyssenkrupp",      # OTC (ADR) — a veces pide plan de pago
    "SSAAY":  "SSAB",              # OTC (ADR) — a veces pide plan de pago
}
ACERERAS_EEUU = {
    "NUE":   "Nucor",
    "STLD":  "Steel Dynamics",
    "CLF":   "Cleveland-Cliffs",
    "CMC":   "Commercial Metals",
    "RS":    "Reliance",
    "WOR":   "Worthington",
    "ATI":   "ATI",
    "CRS":   "Carpenter Technology",
    # "X":   "U.S. Steel",  # la deslistaron el 18-jun-2025 (la compró Nippon); ya no cotiza
}

# Acereras vía Yahoo (gratis y sí cubre bolsas de fuera que el plan gratis de
# Twelve Data no me da). símbolo Yahoo -> (nombre, moneda en que cotiza).
# La moneda la uso para pasar el precio a USD (par USD<moneda>=X de Yahoo), así
# las tarjetas en "$" no mienten y el índice compara todo en una sola moneda.
# Estas se suman al grupo "mundial". A Rusia (Severstal, NLMK, MMK) la dejo fuera
# a propósito: con las sanciones, lo que da Yahoo de la bolsa de Moscú no es
# confiable. Si un símbolo no responde, lo brinco y va a "fallaron".
ACERERAS_YAHOO = {
    # China — el #1 del mundo; las matrices son estatales y cotizan vía filiales
    "600019.SS":     ("Baoshan Iron & Steel",  "CNY"),  # el brazo bursátil de Baowu (#1 mundial)
    "0347.HK":       ("Angang Steel",           "HKD"),  # grupo Ansteel
    "0323.HK":       ("Maanshan Iron & Steel",  "HKD"),
    "000709.SZ":     ("HBIS",                    "CNY"),
    "000959.SZ":     ("Shougang",                "CNY"),
    # India — el #2 del mundo
    "TATASTEEL.NS":  ("Tata Steel",   "INR"),
    "JSWSTEEL.NS":   ("JSW Steel",    "INR"),
    "SAIL.NS":       ("SAIL",         "INR"),
    "JINDALSTEL.NS": ("Jindal Steel", "INR"),
}

# Small-caps de EE.UU. que el plan GRATIS de Twelve Data NO da (responde 404:
# "este símbolo es del plan Grow/Venture"). Las jalo por Yahoo, que sí las cubre
# gratis. Cotizan en USD, así que obtener_yahoo no hace conversión. Importante:
# estas se suman al grupo EE.UU. (no al mundial), que es donde corresponden.
ACERERAS_EEUU_YAHOO = {
    "MTUS": ("Metallus",            "USD"),  # acero especial (ex-TimkenSteel)
    "IIIN": ("Insteel Industries",  "USD"),
    "FRD":  ("Friedman Industries", "USD"),  # NYSE American
}

# El país de cada acerera, para etiquetar las tarjetas y el selector del panel
# (en lugar del genérico "Mundial"). Si una empresa no está aquí, uso su grupo
# grupo como respaldo. Lo edito a gusto: Ternium, por ejemplo, está domiciliada
# en Luxemburgo pero opera sobre todo en México/Argentina (grupo Techint), así
# que la puedo poner como "México" si quiero.
PAIS_ACERERA = {
    "ArcelorMittal":        "Luxemburgo",
    "Gerdau":               "Brasil",
    "Nippon Steel":         "Japón",
    "POSCO":                "Corea del Sur",
    "SSAB":                 "Suecia",
    "Ternium":              "Argentina",
    "thyssenkrupp":         "Alemania",
    "Tata Steel":           "India",
    "JSW Steel":            "India",
    "SAIL":                 "India",
    "Jindal Steel":         "India",
    "Baoshan Iron & Steel": "China",
    "Angang Steel":         "China",
    "Maanshan Iron & Steel": "China",
    "HBIS":                 "China",
    "Shougang":             "China",
    "Nucor":                "EE. UU.",
    "Steel Dynamics":       "EE. UU.",
    "Cleveland-Cliffs":     "EE. UU.",
    "U.S. Steel":           "EE. UU.",
    "Commercial Metals":    "EE. UU.",
    "Reliance":             "EE. UU.",
    "Worthington":          "EE. UU.",
    "ATI":                  "EE. UU.",
    "Carpenter Technology": "EE. UU.",
    "Insteel Industries":   "EE. UU.",
    "Metallus":             "EE. UU.",
    "Friedman Industries":  "EE. UU.",
}

# A cada divisa le agrego también la columna inversa "USD por <ISO>"
# (cuántos dólares vale 1 unidad de esa moneda = 1 / "X por USD").
# Uso 6 decimales porque las inversas van desde ~1.16 (EUR) hasta ~0.0008 (KRW).
FMT_INV = "#,##0.000000"


def _fx_columns():
    """(nombre, formato) de las columnas de divisas: la directa + la inversa, en orden."""
    cols = []
    for iso, (label, fmt) in MONEDAS.items():
        cols.append((label, fmt))
        cols.append((f"USD por {iso}", FMT_INV))
    return cols


def log(msg):
    linea = f"{dt.datetime.now():%Y-%m-%d %H:%M:%S}  {msg}"
    print(linea)
    with open(LOG, "a", encoding="utf-8") as f:
        f.write(linea + "\n")


# ------------------------------------------------------------------
# 1) Descarga de datos
# ------------------------------------------------------------------
def obtener_henry_hub():
    """
    Henry Hub diario. Lo principal lo saco del futuro front-month NG=F por Yahoo
    (yfinance). Lo prefiero sobre el spot de EIA porque es un precio forward
    financiero: sale mucho más suave, sin los brincos del mercado físico de
    contado. Si Yahoo se cae (a veces bloquea peticiones automáticas en CI), me
    voy de respaldo al spot de EIA (NG.RNGWHHD.D) para nunca quedarme sin Henry
    Hub. Dejo el nombre de columna como COL_HH para no romper nada más adelante.
    """
    # 1) Plan A: Yahoo NG=F (futuro front-month, continuo)
    try:
        import yfinance as yf
        raw = yf.download("NG=F", start=FECHA_INICIO, interval="1d",
                          auto_adjust=True, progress=False)
        if (raw is not None and not raw.empty
                and "Close" in raw.columns.get_level_values(0)):
            close = raw["Close"]
            if hasattr(close, "columns"):     # yfinance a veces me da MultiIndex
                close = close.iloc[:, 0]
            df = close.reset_index()
            df.columns = ["fecha", COL_HH]
            df["fecha"] = pd.to_datetime(df["fecha"])
            df[COL_HH] = pd.to_numeric(df[COL_HH], errors="coerce")
            df = df.dropna()
            df = df[df["fecha"] >= pd.Timestamp(FECHA_INICIO)].sort_values("fecha")
            if not df.empty:
                log(f"Henry Hub: Yahoo NG=F (futuro), {len(df)} días")
                return df.reset_index(drop=True)
        log("  Aviso: Yahoo NG=F vino vacío; uso respaldo EIA spot.")
    except Exception as e:
        log(f"  Aviso: Yahoo NG=F falló ({e}); uso respaldo EIA spot.")

    # 2) Plan B: spot diario de EIA (NG.RNGWHHD.D)
    url = "https://api.eia.gov/v2/seriesid/NG.RNGWHHD.D"
    r = requests.get(url, params={"api_key": EIA_API_KEY}, timeout=60)
    r.raise_for_status()
    datos = r.json()["response"]["data"]
    df = pd.DataFrame(datos)[["period", "value"]]
    df.columns = ["fecha", COL_HH]
    df["fecha"] = pd.to_datetime(df["fecha"])
    df[COL_HH] = pd.to_numeric(df[COL_HH], errors="coerce")
    df = df[df["fecha"] >= FECHA_INICIO].sort_values("fecha")
    log(f"Henry Hub: respaldo EIA spot, {len(df)} días")
    return df.reset_index(drop=True)


def obtener_fred(series_id, observation_start=FECHA_INICIO):
    """Una serie diaria/mensual de FRED -> DataFrame[fecha, valor]. Mi función comodín."""
    url = "https://api.stlouisfed.org/fred/series/observations"
    params = {
        "series_id": series_id,
        "api_key": FRED_API_KEY,
        "file_type": "json",
        "observation_start": observation_start,
    }
    r = requests.get(url, params=params, timeout=60)
    r.raise_for_status()
    obs = r.json()["observations"]
    df = pd.DataFrame(obs)[["date", "value"]]
    df.columns = ["fecha", "valor"]
    df["fecha"] = pd.to_datetime(df["fecha"])
    df["valor"] = pd.to_numeric(df["valor"].replace(".", None), errors="coerce")
    return df


def obtener_fx(desde=FECHA_INICIO, reintentos=3):
    """
    Tipo de cambio diario desde Frankfurter (tipos de referencia del BCE).
    Con base=USD me da "unidades por 1 USD" directo, sin invertir nada. Pido
    SOLO las divisas de MONEDAS para que la respuesta sea chica y no se me corte;
    si la conexión se cae, reintenta. (Esto lo blindé porque el API se cayó una
    vez y me tumbó toda la corrida.)
    """
    url = "https://api.frankfurter.dev/v2/rates"
    params = {
        "base": "USD",
        "from": desde,
        "quotes": ",".join(MONEDAS.keys()),   # solo mis divisas (param de la v2)
    }
    ultimo_error = None
    for intento in range(1, reintentos + 1):
        try:
            r = requests.get(url, params=params, timeout=120)
            r.raise_for_status()
            data = r.json()
            break
        except (requests.exceptions.ChunkedEncodingError,
                requests.exceptions.ConnectionError,
                requests.exceptions.Timeout) as e:
            ultimo_error = e
            log(f"  Frankfurter falló (intento {intento}/{reintentos}): {e}")
            if intento == reintentos:
                raise
    else:
        raise ultimo_error

    # Frankfurter v2 me devuelve una lista de registros {date, base, quote, rate}.
    # Dejo el parser tolerante por si el rango llega como dict anidado.
    if isinstance(data, list):
        df = pd.DataFrame(data)
    elif isinstance(data, dict) and "rates" in data:
        rates = data["rates"]
        registros = []
        if rates and all(isinstance(v, dict) for v in rates.values()):
            for fecha, d in rates.items():
                for cur, val in d.items():
                    registros.append({"date": fecha, "quote": cur, "rate": val})
        else:
            for cur, val in rates.items():
                registros.append({"date": data.get("date"), "quote": cur, "rate": val})
        df = pd.DataFrame(registros)
    else:
        raise ValueError("Respuesta de Frankfurter con formato inesperado")

    df = df[df["quote"].isin(MONEDAS)].copy()
    df["fecha"] = pd.to_datetime(df["date"])
    df["rate"] = pd.to_numeric(df["rate"], errors="coerce")
    wide = df.pivot_table(index="fecha", columns="quote", values="rate").reset_index()
    wide = wide.rename(columns={c: MONEDAS[c][0] for c in MONEDAS if c in wide.columns})
    wide.columns.name = None

    # le pego la columna inversa (USD por X) junto a cada divisa y ordeno
    orden = ["fecha"]
    for iso, (label, _fmt) in MONEDAS.items():
        if label in wide.columns:
            inv = f"USD por {iso}"
            wide[inv] = 1.0 / wide[label]
            orden += [label, inv]
    return wide[orden]


def obtener_inflacion():
    """CPI mensual (CPIAUCSL) + la variación interanual."""
    # me echo 13 meses para atrás del inicio para poder calcular el YoY de enero-2021
    df = obtener_fred("CPIAUCSL", observation_start="2019-12-01")
    df = df.rename(columns={"valor": "CPI"}).sort_values("fecha")
    df["Inflacion YoY"] = df["CPI"].pct_change(12)   # fracción (0.034 = 3.4%)
    df = df[df["fecha"] >= FECHA_INICIO].reset_index(drop=True)
    return df


def obtener_macro():
    """
    Macro de EE. UU. (FRED), todo en una sola tabla mensual:
      - Inflación YoY (CPI)            fracción (0.03 = 3%)
      - Desempleo                      % (UNRATE), mensual
      - PIB real (crec. anualizado)    % (A191RL1Q225SBEA), trimestral
    El PIB es trimestral, así que solo trae dato en meses de fin de trimestre;
    el resto de filas quedan en blanco y eso es normal, no es un bug.
    """
    infl = obtener_inflacion()[["fecha", "CPI", "Inflacion YoY"]]

    log("Descargando desempleo (FRED:UNRATE)...")
    des = obtener_fred("UNRATE").rename(columns={"valor": "Desempleo"})

    log("Descargando PIB (FRED:A191RL1Q225SBEA)...")
    pib = obtener_fred("A191RL1Q225SBEA").rename(columns={"valor": "PIB crec."})

    macro = infl.merge(des, on="fecha", how="outer").merge(pib, on="fecha", how="outer")
    macro = macro[macro["fecha"] >= FECHA_INICIO].sort_values("fecha").reset_index(drop=True)
    return macro


def obtener_construccion():
    """
    Construcción de EE. UU. (FRED, datos del Census), mensual. Para mí son
    indicadores líderes de demanda de acero:
      - Housing Starts        viviendas iniciadas, miles de unidades, SAAR (HOUST)
      - Construction Spending gasto total en construcción, SAAR (TTLCONS).
                              FRED me lo da en MILLONES de USD; aquí lo paso a
                              MILES DE MILLONES para que la cifra se lea bien.
    Devuelve [fecha, Housing Starts, Construction Spending] o None.
    """
    log("Descargando Housing Starts (FRED:HOUST)...")
    hs = obtener_fred("HOUST").rename(columns={"valor": "Housing Starts"})

    log("Descargando Construction Spending (FRED:TTLCONS)...")
    cs = obtener_fred("TTLCONS").rename(columns={"valor": "Construction Spending"})
    cs["Construction Spending"] = cs["Construction Spending"] / 1000.0  # de millones a miles de millones

    con = hs.merge(cs, on="fecha", how="outer")
    con = con[con["fecha"] >= FECHA_INICIO].sort_values("fecha").reset_index(drop=True)
    return con


# Chatarra ferrosa (FRED / BLS, PPI por commodity), mensual.
#   id de FRED -> nombre de columna
# OJO conmigo mismo: esto son ÍNDICES de precio (base 1982=100), NO dólares por
# tonelada. El nivel crudo no se compara entre grados; lo que sirve es el % de
# cambio mensual, que es justo lo que pinto en la tarjeta del panel.
FRED_CHATARRA = {
    "WPU10121191": "HMS (Heavy Melting)",
    "WPU10121192": "Bundles (prime)",
    "WPU10121193": "Shredded",
    "WPU101211":   "Chatarra carbono (agregado)",
}


def obtener_chatarra():
    """
    Índices de precio de chatarra ferrosa de EE. UU. (FRED, PPI del BLS),
    mensual. Es el insumo metálico #1 del melt shop, así que me sirve como
    termómetro de costo. Devuelve [fecha, <grados...>] o None.
    """
    df = None
    for sid, nombre in FRED_CHATARRA.items():
        log(f"Descargando chatarra {nombre} (FRED:{sid})...")
        s = obtener_fred(sid).rename(columns={"valor": nombre})
        df = s if df is None else df.merge(s, on="fecha", how="outer")
    if df is None:
        return None
    df = df[df["fecha"] >= FECHA_INICIO].sort_values("fecha").reset_index(drop=True)
    return df


# ------------------------------------------------------------------
# 1c) Top 10 países productores de acero (worldsteel)
# ------------------------------------------------------------------
# worldsteel no tiene API pública, así que esto raspa el HTML de sus press
# releases mensuales. La Tabla 2 ("Top 10 steel-producing countries") siempre
# trae la misma estructura, así que es estable. Si un día cambian el formato,
# revisar _parsear_top10(). Si todo falla devuelve None y la sección se omite,
# igual que chatarra/construcción.
_WS_HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                   "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"),
    "Accept-Language": "en-US,en;q=0.9",
}
_WS_MESES = {"january": 1, "february": 2, "march": 3, "april": 4, "may": 5,
             "june": 6, "july": 7, "august": 8, "september": 9, "october": 10,
             "november": 11, "december": 12}


def _ws_get(url, intentos=4, espera_base=2):
    """GET con backoff exponencial (2,4,8,16s), mismo espíritu que obtener_fx."""
    import time as _t
    ultimo = None
    for i in range(intentos):
        try:
            r = requests.get(url, headers=_WS_HEADERS, timeout=30)
            r.raise_for_status()
            return r.text
        except requests.RequestException as e:
            ultimo = e
            espera = espera_base * (2 ** i)
            log(f"  worldsteel intento {i+1}/{intentos} falló ({e}); reintento en {espera}s")
            _t.sleep(espera)
    log(f"  worldsteel se rindió tras {intentos} intentos: {ultimo}")
    return None


def _ws_ultimo_comunicado(html):
    """Del índice del año saca la URL del comunicado de producción más reciente.
    Ordeno por la fecha del dato (mes-año del slug), no por el orden del HTML."""
    import re
    patron = re.compile(r"press-releases/(\d{4})/([a-z]+)-(\d{4})-crude-steel-production[^\"']*")
    enc = {}
    for m in re.finditer(patron, html):
        mes_txt, anio = m.group(2), int(m.group(3))
        if mes_txt not in _WS_MESES:
            continue
        url = "https://worldsteel.org/media/" + m.group(0)
        if not url.endswith("/"):
            url += "/"
        enc[(anio, _WS_MESES[mes_txt])] = url
    if not enc:
        return None, None
    clave = max(enc)
    return enc[clave], clave


def _ws_parsear_top10(html):
    """Saca la Tabla 2 (Top 10). La identifico porque alguna fila arranca con
    China (la tabla de regiones empieza con Africa). Parseo con lxml directo,
    SIN pandas.read_html, porque read_html se cae a html5lib cuando lxml tropieza
    y ese paquete no está en el runner. Con lxml.html (libxml2) basta y sobra."""
    import re
    from lxml import html as lh

    def _num(txt):
        # normalizo el número: quito separadores de miles (coma o espacio),
        # el espacio duro (nbsp) y el menos "bonito" Unicode que a veces usan
        t = (txt.replace(",", "").replace("\u00a0", "").replace(" ", "")
                .replace("\u2212", "-"))
        return float(t)

    try:
        doc = lh.fromstring(html)
    except Exception as e:
        log(f"  worldsteel: no pude parsear el HTML ({e})")
        return None

    for tabla in doc.xpath("//table"):
        filas = []
        for tr in tabla.xpath(".//tr"):
            celdas = [c.text_content().strip() for c in tr.xpath("./td | ./th")]
            if celdas:
                filas.append(celdas)
        # ¿es la del Top 10? alguna fila debe arrancar con China
        if not any(len(r) >= 5 and "china" in r[0].lower() for r in filas):
            continue
        paises = []
        for r in filas:
            if len(r) < 5:
                continue
            try:
                fila = {
                    "pais": re.sub(r"\s*\(e\)", "", r[0]).strip(),  # quito el (e) de estimado
                    "mes_mt": _num(r[1]),
                    "var_mes_pct": _num(r[2]),
                    "ytd_mt": _num(r[3]),
                    "var_ytd_pct": _num(r[4]),
                }
            except ValueError:
                continue  # encabezado u otra fila no numérica, la salto
            paises.append(fila)
            if len(paises) >= 10:
                break
        return paises or None
    log("  worldsteel: ninguna tabla con China; quizá cambió el formato")
    return None


# Respaldo embebido. Si worldsteel bloquea al runner (como me pasó con Stooq),
# uso estos snapshots reales para que la pestaña nunca quede vacía. Se refrescan
# solos en cuanto el scraping vuelva a funcionar. Para actualizar a mano: copio
# la Tabla 2 del comunicado de cada mes en https://worldsteel.org/media/press-releases/
_WS_FALLBACK_MESES = [
    {"periodo": "2026-05",
     "fuente_url": "https://worldsteel.org/media/press-releases/2026/may-2026-crude-steel-production/",
     "paises": [
        {"pais": "China",         "mes_mt": 84.4, "var_mes_pct": -2.7, "ytd_mt": 415.5, "var_ytd_pct": -3.9},
        {"pais": "India",         "mes_mt": 14.1, "var_mes_pct":  1.9, "ytd_mt":  72.9, "var_ytd_pct":  7.8},
        {"pais": "United States", "mes_mt":  7.5, "var_mes_pct":  9.2, "ytd_mt":  35.6, "var_ytd_pct":  6.8},
        {"pais": "Japan",         "mes_mt":  7.0, "var_mes_pct":  1.7, "ytd_mt":  33.6, "var_ytd_pct": -0.7},
        {"pais": "South Korea",   "mes_mt":  5.4, "var_mes_pct":  3.3, "ytd_mt":  26.4, "var_ytd_pct":  2.7},
        {"pais": "Russia",        "mes_mt":  5.6, "var_mes_pct": -5.4, "ytd_mt":  26.4, "var_ytd_pct": -10.0},
        {"pais": "Türkiye",       "mes_mt":  3.4, "var_mes_pct":  8.9, "ytd_mt":  16.5, "var_ytd_pct":  6.8},
        {"pais": "Germany",       "mes_mt":  3.2, "var_mes_pct":  7.3, "ytd_mt":  15.7, "var_ytd_pct":  8.8},
        {"pais": "Brazil",        "mes_mt":  2.8, "var_mes_pct":  2.4, "ytd_mt":  13.4, "var_ytd_pct": -1.9},
        {"pais": "Viet Nam",      "mes_mt":  2.6, "var_mes_pct": 27.2, "ytd_mt":  12.6, "var_ytd_pct": 26.8},
     ]},
    {"periodo": "2026-04",
     "fuente_url": "https://worldsteel.org/media/press-releases/2026/april-2026-crude-steel-production/",
     "paises": [
        {"pais": "China",         "mes_mt": 83.6, "var_mes_pct": -2.8, "ytd_mt": 331.1, "var_ytd_pct": -4.1},
        {"pais": "India",         "mes_mt": 13.8, "var_mes_pct":  3.9, "ytd_mt":  58.7, "var_ytd_pct":  9.4},
        {"pais": "United States", "mes_mt":  7.2, "var_mes_pct":  9.4, "ytd_mt":  28.1, "var_ytd_pct":  6.6},
        {"pais": "Japan",         "mes_mt":  6.6, "var_mes_pct":  0.3, "ytd_mt":  26.7, "var_ytd_pct": -1.2},
        {"pais": "South Korea",   "mes_mt":  5.2, "var_mes_pct":  4.8, "ytd_mt":  21.0, "var_ytd_pct":  2.5},
        {"pais": "Russia",        "mes_mt":  5.0, "var_mes_pct": -12.4, "ytd_mt": 20.6, "var_ytd_pct": -12.0},
        {"pais": "Türkiye",       "mes_mt":  3.3, "var_mes_pct":  9.4, "ytd_mt":  13.0, "var_ytd_pct":  6.3},
        {"pais": "Germany",       "mes_mt":  3.2, "var_mes_pct":  9.5, "ytd_mt":  12.5, "var_ytd_pct":  9.1},
        {"pais": "Brazil",        "mes_mt":  2.7, "var_mes_pct":  2.8, "ytd_mt":  10.8, "var_ytd_pct": -1.6},
        {"pais": "Viet Nam",      "mes_mt":  2.1, "var_mes_pct":  4.0, "ytd_mt":   8.5, "var_ytd_pct":  8.4},
     ]},
]
_WS_FALLBACK = {
    "periodo": _WS_FALLBACK_MESES[0]["periodo"],
    "fuente_url": _WS_FALLBACK_MESES[0]["fuente_url"],
    "respaldo": True,   # bandera: este dato vino del snapshot, no de la corrida en vivo
    "paises": _WS_FALLBACK_MESES[0]["paises"],
    "meses": _WS_FALLBACK_MESES,
}


def _ws_meses_recientes(n=6):
    """Devuelve los últimos n (año, mes) de datos, del más reciente al más viejo.
    El comunicado de un mes sale ~3-4 semanas después, así que arranco en el mes
    pasado respecto a hoy."""
    hoy = dt.date.today()
    y, m = hoy.year, hoy.month
    salida = []
    for _ in range(n):
        m -= 1
        if m == 0:
            m = 12
            y -= 1
        salida.append((y, m))
    return salida


def _ws_url_directa(anio_dato, mes_dato):
    """Construyo la URL del comunicado mensual directo, sin pasar por el índice.
    El año de la ruta es el de PUBLICACIÓN (el mes siguiente), que para ene-nov es
    el mismo año del dato y para diciembre es el año siguiente."""
    nombres = {1: "january", 2: "february", 3: "march", 4: "april", 5: "may",
               6: "june", 7: "july", 8: "august", 9: "september", 10: "october",
               11: "november", 12: "december"}
    anio_pub = anio_dato if mes_dato != 12 else anio_dato + 1
    return (f"https://worldsteel.org/media/press-releases/{anio_pub}/"
            f"{nombres[mes_dato]}-{anio_dato}-crude-steel-production/")


def _ws_valido(top10):
    """Sanity check: que sea una lista decente con China adentro."""
    return bool(top10) and len(top10) >= 8 and any(
        "china" in p["pais"].lower() for p in top10)


def _ws_diagnostico(html, url):
    """Cuando bajé algo pero no salió la tabla, dejo pistas en la bitácora para
    saber si fue bloqueo de IP (muro de Cloudflare/consentimiento) o cambio de
    formato. Esto es lo que me va a decir la causa real."""
    h = (html or "").lower()
    n_tablas = h.count("<table")
    marcas = [m for m in ("just a moment", "cloudflare", "captcha", "access denied",
                          "are you human", "enable javascript", "cookie")
              if m in h]
    log(f"  worldsteel DIAG [{url}]: {len(html or '')} chars, "
        f"{n_tablas} <table>, China={'sí' if 'china' in h else 'NO'}, "
        f"sospechas={marcas or 'ninguna'}")


def obtener_productores(n_meses=8):
    """Top 10 países productores de acero (worldsteel), de los últimos n_meses.
    Intenta URL directa mes por mes; si worldsteel bloquea al runner, cae al
    respaldo embebido. Nunca devuelve None: la pestaña siempre muestra algo.

    Estructura: {periodo, fuente_url, paises (mes más reciente), meses:[...]}.
    'meses' va del más reciente al más viejo; cada uno trae su propio Top 10."""
    meses = []
    primero = True
    for (y, m) in _ws_meses_recientes(n_meses):
        url = _ws_url_directa(y, m)
        html = _ws_get(url, intentos=2)        # pocos reintentos: si 404 paso al siguiente
        if html:
            top = _ws_parsear_top10(html)
            if _ws_valido(top):
                meses.append({"periodo": f"{y}-{m:02d}", "fuente_url": url, "paises": top})
            elif primero:
                # bajé algo pero no sirvió en el mes más reciente: dejo diagnóstico
                _ws_diagnostico(html, url)
        primero = False

    # respaldo por índice si la URL directa no trajo nada (por si cambió el patrón)
    if not meses:
        anio = dt.date.today().year
        for a in (anio, anio - 1):
            idx = _ws_get(f"https://worldsteel.org/media/press-releases/{a}/")
            if not idx:
                continue
            url_pr, clave = _ws_ultimo_comunicado(idx)
            if not url_pr:
                continue
            html_pr = _ws_get(url_pr)
            top = _ws_parsear_top10(html_pr) if html_pr else None
            if _ws_valido(top):
                ad, md = clave
                meses.append({"periodo": f"{ad}-{md:02d}", "fuente_url": url_pr, "paises": top})
            break

    if meses:
        log(f"Productores: worldsteel, {len(meses)} mes(es) "
            f"({meses[-1]['periodo']} a {meses[0]['periodo']})")
        reciente = meses[0]
        return {
            "periodo": reciente["periodo"],
            "fuente_url": reciente["fuente_url"],
            "paises": reciente["paises"],
            "meses": meses,
        }

    # nada jaló (probable bloqueo de IP): respaldo embebido para no dejar vacío
    log(f"  worldsteel no accesible desde el runner; uso respaldo embebido "
        f"({len(_WS_FALLBACK['meses'])} meses). Revisa el DIAG de arriba para la causa.")
    return _WS_FALLBACK


# ------------------------------------------------------------------
# 1b) Acereras: precios de acciones (Twelve Data + Yahoo) -> índices y correlación
# ------------------------------------------------------------------
def obtener_twelvedata(symbol, reintentos=4):
    """
    Cierre diario de una acción desde Twelve Data (JSON, con llave).
    Me devuelve una Serie indexada por fecha, o None. Y registra POR QUÉ falló,
    que eso me ahorró horas de debug.

    Lo que tengo que recordar del plan GRATIS de Twelve Data:
      - Frecuencia: 8 peticiones por minuto (el throttle real lo manejo en
        'jalar', esperando entre símbolo y símbolo).
      - Tope diario: 800 créditos; cada llamada de aquí gasta 1.
      - Si pego con el tope (code 429 / "credits"), espero ~65 s y reintento.
      - Si el símbolo no existe o el plan no lo cubre, no insisto: lo brinco.
    Twelve Data me devuelve HTTP 200 hasta cuando hay error lógico; el detalle
    real viene en el campo 'status'/'code' del JSON, así que eso es lo que reviso.
    """
    if not TWELVEDATA_API_KEY:
        log(f"  TwelveData {symbol}: falta TWELVEDATA_API_KEY (Secret del repo / .env)")
        return None

    url = "https://api.twelvedata.com/time_series"
    params = {
        "symbol":     symbol,
        "interval":   "1day",
        "start_date": FECHA_INICIO,
        "outputsize": 5000,          # de 2021 a hoy me sobra con esto
        "order":      "ASC",
        "apikey":     TWELVEDATA_API_KEY,
        "format":     "JSON",
    }
    for intento in range(1, reintentos + 1):
        try:
            r = requests.get(url, params=params, timeout=60)
            try:
                data = r.json()
            except ValueError:
                log(f"  TwelveData {symbol}: respuesta no-JSON (status {r.status_code}) "
                    f"-> '{(r.text or '')[:60]}'")
                time.sleep(3.0 * intento)
                continue

            # Error lógico: {"status":"error","code":...,"message":...}
            if isinstance(data, dict) and data.get("status") == "error":
                code = data.get("code")
                msg = (data.get("message") or "")[:90]
                if code == 429 or "credit" in msg.lower() or "limit" in msg.lower():
                    espera = 65
                    log(f"  TwelveData {symbol}: LÍMITE (code {code}) -> espero {espera}s -> '{msg}'")
                    time.sleep(espera)
                    continue
                # Símbolo que no existe o que mi plan no cubre: no insisto.
                log(f"  TwelveData {symbol}: ERROR (code {code}) -> '{msg}'")
                return None

            valores = data.get("values") if isinstance(data, dict) else None
            if not valores:
                log(f"  TwelveData {symbol}: sin 'values' -> '{str(data)[:90]}'")
                return None

            df = pd.DataFrame(valores)
            if "datetime" not in df.columns or "close" not in df.columns:
                log(f"  TwelveData {symbol}: columnas raras {list(df.columns)[:5]}")
                return None

            s = pd.Series(pd.to_numeric(df["close"], errors="coerce").values,
                          index=pd.to_datetime(df["datetime"], errors="coerce"))
            s.index.name = "fecha"
            s = s.dropna().sort_index()
            s = s[s.index >= FECHA_INICIO]
            if not s.empty:
                return s
            log(f"  TwelveData {symbol}: 0 filas tras filtrar desde {FECHA_INICIO}")
            return None
        except Exception as e:
            log(f"  TwelveData {symbol} intento {intento}: {e}")
            time.sleep(3.0 * intento)
    return None


_FX_YAHOO = {}   # cache: moneda -> Serie "unidades por 1 USD" (o None)


def _fx_por_usd(moneda):
    """Serie diaria 'unidades de <moneda> por 1 USD' desde Yahoo (par USD<moneda>=X).
    La uso para pasar a USD los precios que cotizan en otra moneda. La cacheo
    para no pedir el mismo par dos veces."""
    if moneda == "USD":
        return None
    if moneda in _FX_YAHOO:
        return _FX_YAHOO[moneda]
    serie = None
    try:
        import yfinance as yf
        raw = yf.download(f"USD{moneda}=X", start=FECHA_INICIO, interval="1d",
                          auto_adjust=True, progress=False)
        if (raw is not None and not raw.empty
                and "Close" in raw.columns.get_level_values(0)):
            close = raw["Close"]
            if hasattr(close, "columns"):
                close = close.iloc[:, 0]
            serie = pd.to_numeric(close, errors="coerce")
            serie.index = pd.to_datetime(serie.index)
            serie = serie.dropna().sort_index()
    except Exception as e:
        log(f"  FX USD{moneda}=X (Yahoo) falló: {e}")
    _FX_YAHOO[moneda] = serie
    return serie


def obtener_yahoo(symbol, moneda="USD", reintentos=3):
    """Cierre diario de una acción desde Yahoo (yfinance), YA PASADO A USD si
    'moneda' no es USD. Cubre Shanghái (.SS), Shenzhen (.SZ), Hong Kong (.HK) e
    India (.NS). Si el símbolo no contesta, devuelvo None y el índice lo aguanta."""
    import yfinance as yf
    for intento in range(1, reintentos + 1):
        try:
            raw = yf.download(symbol, start=FECHA_INICIO, interval="1d",
                              auto_adjust=True, progress=False)
            if (raw is None or raw.empty
                    or "Close" not in raw.columns.get_level_values(0)):
                log(f"  Yahoo {symbol}: vacío (intento {intento}/{reintentos})")
                time.sleep(2.0 * intento)
                continue
            close = raw["Close"]
            if hasattr(close, "columns"):
                close = close.iloc[:, 0]
            s = pd.to_numeric(close, errors="coerce")
            s.index = pd.to_datetime(s.index)
            s = s.dropna().sort_index()
            s = s[s.index >= pd.Timestamp(FECHA_INICIO)]
            if s.empty:
                log(f"  Yahoo {symbol}: 0 filas tras filtrar desde {FECHA_INICIO}")
                return None
            if moneda != "USD":                      # lo paso a USD
                fx = _fx_por_usd(moneda)             # unidades por 1 USD
                if fx is None or fx.empty:
                    log(f"  Yahoo {symbol}: sin FX {moneda}; lo dejo en moneda local")
                else:
                    fx_al = fx.reindex(s.index, method="ffill")
                    s = (s / fx_al).dropna()         # local / (local por USD) = USD
            s.index.name = "fecha"
            return s
        except Exception as e:
            log(f"  Yahoo {symbol} intento {intento}: {e}")
            time.sleep(2.0 * intento)
    return None


def _indice_grupo(precios):
    """
    Índice equal-weight (base 100) armado con el rendimiento diario promedio de
    las acciones del grupo. Lo hago así para que entren/salgan empresas sin
    saltos artificiales. Recorto los movimientos diarios a ±50% para que un split
    no me ensucie el índice.
    """
    rets = {}
    for tk, s in precios.items():
        s = s.dropna().sort_index()
        if len(s) < 2:
            continue
        rets[tk] = s.pct_change().clip(-0.5, 0.5)
    if not rets:
        return None
    R = pd.DataFrame(rets).sort_index()
    prom = R.mean(axis=1, skipna=True).fillna(0.0)   # rendimiento equal-weight del grupo
    return 100.0 * (1 + prom).cumprod()


def _exportar_precios_acereras(grupos):
    """
    Armo la estructura de PRECIOS por acción para el panel, al estilo de las
    divisas: una tarjeta por empresa (último precio + % de cambio + fecha) y la
    serie completa de cada una para la gráfica con selector.

    'grupos' es una lista de (dic_ticker_a_nombre, precios_ticker_a_serie).
    Devuelve {"fecha": [...], "series": {nombre: [...]}, "resumen": [...]} o None.
    Los precios son el cierre tal cual lo da Twelve Data (USD para las gringas) o
    ya pasado a USD para las de Yahoo. Redondeo a 2.
    """
    cols = {}
    grupo_de = {}
    for dic, precios in grupos:
        etiqueta = "Mundial" if dic is grupos[0][0] else "EE. UU."
        for tk, s in precios.items():
            nombre = dic.get(tk, tk)
            cols[nombre] = s
            grupo_de[nombre] = etiqueta
    if not cols:
        return None

    P = pd.DataFrame(cols).sort_index()
    P = P[P.index >= FECHA_INICIO]
    fechas = [d.strftime("%Y-%m-%d") for d in P.index]

    series, resumen = {}, []
    for nombre in P.columns:
        col = P[nombre]
        series[nombre] = [None if pd.isna(v) else round(float(v), 2) for v in col]
        valida = col.dropna()
        if valida.empty:
            continue
        ult = float(valida.iloc[-1])
        fecha_ult = valida.index[-1].strftime("%Y-%m-%d")
        cambio = None
        if len(valida) >= 2:
            prev = float(valida.iloc[-2])
            if prev != 0:
                cambio = round((ult - prev) / prev * 100.0, 2)
        resumen.append({"nombre": nombre, "grupo": grupo_de.get(nombre, ""),
                        "pais": PAIS_ACERERA.get(nombre, grupo_de.get(nombre, "")),
                        "valor": round(ult, 2), "cambio": cambio,
                        "fecha": fecha_ult})

    # Mundial primero, luego EE. UU., y dentro de cada grupo en orden alfabético.
    resumen.sort(key=lambda x: (x["grupo"] != "Mundial", x["nombre"]))
    return {"fecha": fechas, "series": series, "resumen": resumen}


def construir_acereras():
    """
    Bajo las acereras de cada grupo, armo los dos índices base 100 y saco su
    correlación (global y móvil a 90 días). Devuelve (df, info).

    El grupo "mundial" me mezcla dos fuentes: Twelve Data (los ADRs gringos) y
    Yahoo (China e India, que el plan gratis de Twelve Data no me da).
    """
    if not TWELVEDATA_API_KEY:
        log("  Aviso: falta TWELVEDATA_API_KEY; las acereras de Twelve Data se "
            "me caen, pero las de Yahoo (China, India) sí entran.")

    def jalar(dic, etiqueta):
        precios, ok, fallaron = {}, [], []
        if not TWELVEDATA_API_KEY:
            return precios, ok, [f"{nom} ({tk})" for tk, nom in dic.items()]
        for tk, nom in dic.items():
            log(f"Descargando {nom} ({tk}) [{etiqueta}]...")
            s = obtener_twelvedata(tk)
            if s is None or s.empty:
                fallaron.append(f"{nom} ({tk})")
            else:
                precios[tk] = s
                ok.append(nom)
            time.sleep(8.0)   # plan gratis: máx 8 peticiones/min -> 1 cada 8 s y no me banean
        return precios, ok, fallaron

    def jalar_yahoo(dic, etiqueta="mundial"):
        precios, ok, fallaron, nombres = {}, [], [], {}
        for tk, (nom, moneda) in dic.items():
            log(f"Descargando {nom} ({tk}) [{etiqueta}/Yahoo {moneda}]...")
            s = obtener_yahoo(tk, moneda=moneda)
            nombres[tk] = nom
            if s is None or getattr(s, "empty", True):
                fallaron.append(f"{nom} ({tk})")
            else:
                precios[tk] = s
                ok.append(nom)
            time.sleep(1.5)   # Yahoo no me aprieta como Twelve Data, con esto basta
        return precios, ok, fallaron, nombres

    p_m, ok_m, fail_m       = jalar(ACERERAS_MUNDIAL, "mundial")   # Twelve Data (ADRs)
    p_y, ok_y, fail_y, nm_y = jalar_yahoo(ACERERAS_YAHOO)          # Yahoo (China, India)
    p_u, ok_u, fail_u       = jalar(ACERERAS_EEUU, "EE.UU.")       # Twelve Data (large caps)
    # small-caps de EE.UU. por Yahoo (Twelve Data no las da gratis): MISMO grupo EE.UU.
    p_uy, ok_uy, fail_uy, nm_uy = jalar_yahoo(ACERERAS_EEUU_YAHOO, "EE.UU.")
    p_u    = {**p_u, **p_uy}
    ok_u   = ok_u + ok_uy
    fail_u = fail_u + fail_uy
    NOMBRES_EEUU = {**ACERERAS_EEUU, **nm_uy}   # para el export de precios

    # junto Twelve Data + Yahoo en un solo grupo "mundial"
    NOMBRES_MUNDIAL = {**ACERERAS_MUNDIAL, **nm_y}
    p_m    = {**p_m, **p_y}
    ok_m   = ok_m + ok_y
    fail_m = fail_m + fail_y

    idx_m = _indice_grupo(p_m)
    idx_u = _indice_grupo(p_u)
    if idx_m is None or idx_u is None:
        log("  Aviso: faltan datos de acereras en algún grupo; se omite la sección.")
        return None, {"corr_global": None, "ok_mundial": ok_m, "ok_eeuu": ok_u,
                      "fallaron": fail_m + fail_u}

    df = pd.concat({"Índice mundial": idx_m, "Índice EE. UU.": idx_u}, axis=1)
    df = df.sort_index()
    df = df[df.index >= FECHA_INICIO].dropna(how="all")
    # rebaso ambos a 100 en su primera fecha común para poder compararlos en la gráfica
    df = df.dropna()
    for c in df.columns:
        df[c] = df[c] / df[c].iloc[0] * 100.0

    # Correlación sobre rendimientos SEMANALES (cierre de viernes). Le doy semana
    # para que todas las bolsas (Asia, Europa, EE. UU.) reflejen lo mismo y así
    # me quito el desfase de husos horarios que ensucia la versión diaria.
    sem = df.resample("W-FRI").last()
    rmw = sem["Índice mundial"].pct_change()
    ruw = sem["Índice EE. UU."].pct_change()
    corr_global = float(rmw.corr(ruw))
    corr_movil = rmw.rolling(13).corr(ruw)            # ~1 trimestre (13 semanas)
    corr_fecha = [d.strftime("%Y-%m-%d") for d in sem.index]
    corr_series = [None if pd.isna(v) else round(float(v), 3) for v in corr_movil]

    # Índice REAL publicado: intento primero el NYSE Arca Steel Index (^STEEL) y,
    # si Yahoo no lo da (pasa seguido, es un índice poco líquido), caigo a SLX
    # (VanEck Steel ETF), que es el ETF que físicamente replica ese índice
    # (~24 acereras globales ponderadas por capitalización) y sí trae precio
    # diario confiable. Cualquiera de los dos se rebasa a 100 igual que mundial/
    # EE. UU. Si ninguno responde, simplemente no sale esa línea (no afecta la
    # correlación, que se calcula solo entre mundial y EE. UU.).
    log("Descargando índice real de acereras [Yahoo]...")
    steel = obtener_yahoo("^STEEL", moneda="USD")
    fuente_indice_real = "^STEEL"
    NOMBRE_INDICE_REAL = "NYSE Arca Steel Index"
    if steel is None or getattr(steel, "empty", True) or steel.reindex(df.index).dropna().shape[0] <= 1:
        log("  ^STEEL: sin datos útiles; probando SLX (ETF) como respaldo...")
        steel = obtener_yahoo("SLX", moneda="USD")
        fuente_indice_real = "SLX"
        NOMBRE_INDICE_REAL = "NYSE Arca Steel Index (SLX ETF)"

    if steel is not None and not getattr(steel, "empty", True):
        s = steel.reindex(df.index).astype(float)
        valida = s.dropna()
        if len(valida) > 1:
            df[NOMBRE_INDICE_REAL] = s / valida.iloc[0] * 100.0
            log(f"  {fuente_indice_real}: {len(valida)} cierres OK")
        else:
            log(f"  {fuente_indice_real}: sin datos en el rango; se omite la línea")
    else:
        log(f"  {fuente_indice_real}: vino vacío; se omite la línea")

    df = df.reset_index().rename(columns={"index": "fecha"})
    if "fecha" not in df.columns:        # por si el índice no se llamó 'fecha'
        df = df.rename(columns={df.columns[0]: "fecha"})

    # Precios individuales por acción (tarjetas + gráfica con selector, igual que las divisas).
    precios_export = _exportar_precios_acereras(
        [(NOMBRES_MUNDIAL, p_m), (NOMBRES_EEUU, p_u)])

    info = {"corr_global": corr_global, "corr_fecha": corr_fecha,
            "corr_series": corr_series, "ok_mundial": ok_m, "ok_eeuu": ok_u,
            "fallaron": fail_m + fail_u, "precios": precios_export}
    log(f"Correlación semanal acereras mundial vs. EE. UU. (desde {FECHA_INICIO}): {corr_global:.2f}")
    return df, info


def construir_diario():
    """Junto todas las series diarias en una sola tabla por fecha."""
    log("Descargando Henry Hub (EIA)...")
    diario = obtener_henry_hub()

    for sid, (nombre, _fmt) in FRED_DIARIAS.items():
        log(f"Descargando {nombre} (FRED:{sid})...")
        s = obtener_fred(sid).rename(columns={"valor": nombre})
        diario = diario.merge(s, on="fecha", how="outer")

    log("Descargando divisas (Frankfurter/BCE)...")
    fx = obtener_fx()
    diario = diario.merge(fx, on="fecha", how="outer")

    diario = diario[diario["fecha"] >= FECHA_INICIO].sort_values("fecha")
    diario = diario.reset_index(drop=True)

    # le quito los picos falsos de la fuente a los precios de energía
    for col in RANGOS_VALIDOS:
        diario = sanear(diario, col)

    return diario


# ------------------------------------------------------------------
# 2) El Excel con formato bonito
# ------------------------------------------------------------------
# Paleta (la misma del panel: carbón, naranja, rojo, dorado, franja cálida)
AZUL   = "EA580C"   # fila de encabezados (naranja) — el nombre quedó de antes, ya sé
AZUL2  = "26262B"   # barra de título (carbón)
ORO    = "FFC400"   # texto de títulos / acentos (dorado)
GRIS   = "FDF1E7"   # franjas alternas (cálido claro)
BLANCO = "FFFFFF"
borde_fino = Border(*(Side(style="thin", color="EADFD3"),) * 4)


def _formatos_diario():
    """Mapa nombre_de_columna -> formato numérico, para la hoja Diario."""
    fmt = {COL_HH: FMT_HH}
    fmt.update({n: f for (n, f) in FRED_DIARIAS.values()})
    fmt.update({n: f for (n, f) in _fx_columns()})
    return fmt


def _series_resumen():
    """Lista ordenada (nombre, formato) para la hoja Resumen."""
    s = [(COL_HH, FMT_HH)]
    s += [(n, f) for (n, f) in FRED_DIARIAS.values()]
    s += _fx_columns()
    return s


def _estilizar_hoja(ws, df, formatos, titulo, col_fecha="fecha"):
    """Vuelco un DataFrame con encabezado, franjas, filtros y formato. Aquí está
    todo el rollo cosmético del Excel."""
    cols = list(df.columns)
    n_col = len(cols)

    # ---- fila de título (fila 1) ----
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_col)
    c = ws.cell(row=1, column=1, value=titulo)
    c.font = Font(name="Calibri", size=14, bold=True, color=ORO)
    c.fill = PatternFill("solid", fgColor=AZUL2)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    # ---- subtítulo con la fecha de generación (fila 2) ----
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_col)
    c = ws.cell(row=2, column=1,
                value=f"Generado: {dt.datetime.now():%Y-%m-%d %H:%M}  ·  "
                      f"Fuentes: EIA, FRED y Frankfurter (BCE)  ·  Inicio: {FECHA_INICIO}")
    c.font = Font(name="Calibri", size=9, italic=True, color="808080")
    c.alignment = Alignment(horizontal="left", indent=1)

    fila_enc = 3
    # ---- encabezados ----
    for j, nombre in enumerate(cols, start=1):
        cell = ws.cell(row=fila_enc, column=j, value=nombre)
        cell.font = Font(name="Calibri", size=10, bold=True, color=BLANCO)
        cell.fill = PatternFill("solid", fgColor=AZUL)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = borde_fino
    ws.row_dimensions[fila_enc].height = 30

    # ---- datos ----
    for i, (_, fila) in enumerate(df.iterrows()):
        r = fila_enc + 1 + i
        franja = PatternFill("solid", fgColor=GRIS) if i % 2 else None
        for j, nombre in enumerate(cols, start=1):
            val = fila[nombre]
            if pd.isna(val):
                val = None
            cell = ws.cell(row=r, column=j, value=val)
            cell.border = borde_fino
            if franja:
                cell.fill = franja
            if nombre == col_fecha:
                cell.number_format = "yyyy-mm-dd"
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.number_format = formatos.get(nombre, "#,##0.00")
                cell.alignment = Alignment(horizontal="right")

    # ---- anchos de columna ----
    for j, nombre in enumerate(cols, start=1):
        letra = get_column_letter(j)
        if nombre == col_fecha:
            ws.column_dimensions[letra].width = 12
        else:
            ws.column_dimensions[letra].width = max(13, min(len(nombre) + 2, 20))

    # ---- congelo encabezado + primera columna, y pongo autofiltro ----
    ws.freeze_panes = ws.cell(row=fila_enc + 1, column=2)
    ultima = get_column_letter(n_col)
    ws.auto_filter.ref = f"A{fila_enc}:{ultima}{fila_enc + len(df)}"


def _hoja_resumen(ws, diario, infl):
    """La hoja de portada con el último valor de cada serie."""
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:C1")
    c = ws.cell(row=1, column=1, value="Resumen de mercados")
    c.font = Font(name="Calibri", size=16, bold=True, color=ORO)
    c.fill = PatternFill("solid", fgColor=AZUL2)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:C2")
    c = ws.cell(row=2, column=1,
                value=f"Último dato disponible · Generado {dt.datetime.now():%Y-%m-%d %H:%M}")
    c.font = Font(size=9, italic=True, color="808080")
    c.alignment = Alignment(indent=1)

    enc = ["Serie", "Último valor", "Fecha del dato"]
    for j, t in enumerate(enc, start=1):
        cell = ws.cell(row=4, column=j, value=t)
        cell.font = Font(bold=True, color=BLANCO)
        cell.fill = PatternFill("solid", fgColor=AZUL)
        cell.alignment = Alignment(horizontal="center")
        cell.border = borde_fino

    r = 5
    for nombre, fmt in _series_resumen():
        if nombre not in diario.columns:
            continue
        sub = diario[["fecha", nombre]].dropna()
        if sub.empty:
            continue
        ult = sub.iloc[-1]
        franja = PatternFill("solid", fgColor=GRIS) if (r % 2 == 0) else None
        v1 = ws.cell(row=r, column=1, value=nombre)
        v2 = ws.cell(row=r, column=2, value=float(ult[nombre]))
        v3 = ws.cell(row=r, column=3, value=ult["fecha"].to_pydatetime())
        v2.number_format = fmt
        v3.number_format = "yyyy-mm-dd"
        for cell in (v1, v2, v3):
            cell.border = borde_fino
            if franja:
                cell.fill = franja
        v1.alignment = Alignment(horizontal="left")
        v2.alignment = Alignment(horizontal="right")
        v3.alignment = Alignment(horizontal="center")
        r += 1

    # macro mensual/trimestral (esto va aparte porque no es diario)
    macro = infl
    filas_macro = [
        ("Inflación YoY (CPI)", "Inflacion YoY", "0.0%"),
        ("Desempleo (%)",       "Desempleo",     '0.0"%"'),
        ("PIB real (crec. anual.)", "PIB crec.",  '0.0"%"'),
    ]
    for etiqueta, col, fmt in filas_macro:
        if col not in macro.columns:
            continue
        sub = macro[["fecha", col]].dropna()
        if sub.empty:
            continue
        ult = sub.iloc[-1]
        franja = PatternFill("solid", fgColor=GRIS) if (r % 2 == 0) else None
        v1 = ws.cell(row=r, column=1, value=etiqueta)
        v2 = ws.cell(row=r, column=2, value=float(ult[col]))
        v3 = ws.cell(row=r, column=3, value=ult["fecha"].to_pydatetime())
        v2.number_format = fmt
        v3.number_format = "yyyy-mm-dd"
        for cell in (v1, v2, v3):
            cell.border = borde_fino
            if franja:
                cell.fill = franja
        v1.alignment = Alignment(horizontal="left")
        v2.alignment = Alignment(horizontal="right")
        v3.alignment = Alignment(horizontal="center")
        r += 1

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16


def _grafico_excel(ws, fila_enc, n_filas, col_idx, titulo, ancla):
    """Mete un mini gráfico de línea nativo de Excel en la hoja Diario."""
    ch = LineChart()
    ch.title = titulo
    ch.height = 7
    ch.width = 16
    ch.style = 2
    datos = Reference(ws, min_col=col_idx, max_col=col_idx,
                      min_row=fila_enc, max_row=fila_enc + n_filas)
    cats = Reference(ws, min_col=1, min_row=fila_enc + 1, max_row=fila_enc + n_filas)
    ch.add_data(datos, titles_from_data=True)
    ch.set_categories(cats)
    ch.legend = None
    ws.add_chart(ch, ancla)


def _hoja_graficos_fx(wb, ws_diario, diario):
    """Una hoja con una gráfica de línea nativa por cada divisa (unidades por 1 USD)."""
    ws = wb.create_sheet("Gráficos FX")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:R1")
    c = ws.cell(row=1, column=1, value="Tipos de cambio  ·  unidades por 1 USD")
    c.font = Font(name="Calibri", size=14, bold=True, color=ORO)
    c.fill = PatternFill("solid", fgColor=AZUL2)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    fila_enc = 3                         # la fila de encabezados en la hoja Diario
    n = len(diario)
    cols = list(diario.columns)
    anclas_col = ["A", "J"]              # dos columnas de gráficas, lado a lado
    paso_fila = 14                       # cuántas filas dejo entre una gráfica y la de abajo
    idx = 0
    for iso, (label, _fmt) in MONEDAS.items():
        if label not in cols:
            continue
        col_idx = cols.index(label) + 1
        ch = LineChart()
        ch.title = label
        ch.height = 6.5
        ch.width = 12
        ch.style = 2
        ch.legend = None
        datos = Reference(ws_diario, min_col=col_idx, max_col=col_idx,
                          min_row=fila_enc, max_row=fila_enc + n)
        cats = Reference(ws_diario, min_col=1,
                         min_row=fila_enc + 1, max_row=fila_enc + n)
        ch.add_data(datos, titles_from_data=True)
        ch.set_categories(cats)
        fila = 3 + (idx // 2) * paso_fila
        ws.add_chart(ch, f"{anclas_col[idx % 2]}{fila}")
        idx += 1


def escribir_excel(diario, infl, ruta, acereras=None):
    wb = Workbook()

    # Hoja 1: Resumen
    ws_res = wb.active
    ws_res.title = "Resumen"
    _hoja_resumen(ws_res, diario, infl)

    # Hoja 2: Diario
    ws_dia = wb.create_sheet("Diario")
    _estilizar_hoja(ws_dia, diario, _formatos_diario(),
                    "Series diarias  ·  Henry Hub, S&P 500, WTI, Brent y divisas")

    # gráficos nativos de Henry Hub y S&P 500
    fila_enc = 3
    n = len(diario)
    cols = list(diario.columns)
    ancla_col = get_column_letter(len(cols) + 2)
    if COL_HH in cols:
        _grafico_excel(ws_dia, fila_enc, n, cols.index(COL_HH) + 1,
                       COL_HH, f"{ancla_col}3")
    if "S&P 500" in cols:
        _grafico_excel(ws_dia, fila_enc, n, cols.index("S&P 500") + 1,
                       "S&P 500", f"{ancla_col}18")

    # Hoja 3: Macro EE. UU.
    ws_mac = wb.create_sheet("Macro EE. UU.")
    _estilizar_hoja(ws_mac, infl,
                    {"CPI": "#,##0.000", "Inflacion YoY": "0.0%",
                     "Desempleo": '0.0"%"', "PIB crec.": '0.0"%"'},
                    "Macro EE. UU.  ·  inflación, desempleo y PIB")

    # Hoja 4: Acereras (los índices; la correlación semanal la dejo en el panel)
    if acereras is not None and not acereras.empty:
        ws_ac = wb.create_sheet("Acereras")
        _estilizar_hoja(ws_ac, acereras,
                        {"Índice mundial": "#,##0.0", "Índice EE. UU.": "#,##0.0"},
                        "Acereras  ·  índice mundial vs. EE. UU. (base 100 = 2021)")

    # Hoja final: una gráfica por divisa
    _hoja_graficos_fx(wb, ws_dia, diario)

    wb.save(ruta)


# ------------------------------------------------------------------
# 3) Los PNG (panel resumen con matplotlib)
# ------------------------------------------------------------------
def crear_graficos(diario, macro, ruta_png):
    fig, ax = plt.subplots(3, 2, figsize=(13, 12))
    fig.suptitle("Mercados y macro — resumen", fontsize=14, fontweight="bold")

    d = diario.set_index("fecha")
    if COL_HH in d:
        ax[0, 0].plot(d.index, d[COL_HH], color="#FF7A18")
        ax[0, 0].set_title(COL_HH)
    for col, c in [("WTI ($/bbl)", "#FF7A18"), ("Brent ($/bbl)", "#E62315")]:
        if col in d:
            ax[0, 1].plot(d.index, d[col], label=col, color=c)
    ax[0, 1].set_title("Petróleo: WTI vs. Brent")
    ax[0, 1].legend(fontsize=8)
    if "S&P 500" in d:
        ax[1, 0].plot(d.index, d["S&P 500"], color="#888c92")
        ax[1, 0].set_title("S&P 500")

    m = macro.set_index("fecha")
    if "Inflacion YoY" in m:
        ai = m["Inflacion YoY"].dropna() * 100
        ax[1, 1].plot(ai.index, ai.values, color="#BF8F00")
        ax[1, 1].set_title("Inflación YoY (%)")
    if "Desempleo" in m:
        de = m["Desempleo"].dropna()
        ax[2, 0].plot(de.index, de.values, color="#2E5496")
        ax[2, 0].set_title("Desempleo (%)")
    if "PIB crec." in m:
        pb = m["PIB crec."].dropna()
        ax[2, 1].plot(pb.index, pb.values, color="#548235", marker="o", markersize=3)
        ax[2, 1].set_title("PIB real — crec. anualizado (%)")

    for a in ax.flat:
        a.grid(True, alpha=0.3)
        a.tick_params(labelsize=8)
    fig.text(0.99, 0.01,
             f"Fuentes: Henry Hub (EIA); S&P 500, WTI, Brent, inflación, "
             f"desempleo y PIB (FRED).  Generado {dt.date.today():%Y-%m-%d}.",
             ha="right", va="bottom", fontsize=7.5, color="#808080")
    fig.tight_layout(rect=[0, 0.02, 1, 0.97])
    fig.savefig(ruta_png, dpi=130)
    plt.close(fig)


def crear_graficos_fx(diario, ruta_png):
    """Un panel con una gráfica por divisa (unidades por 1 USD)."""
    monedas = [label for (label, _f) in MONEDAS.values() if label in diario.columns]
    if not monedas:
        return
    ncols = 2
    nfilas = (len(monedas) + ncols - 1) // ncols
    fig, ax = plt.subplots(nfilas, ncols, figsize=(13, 3 * nfilas))
    fig.suptitle("Tipos de cambio — unidades por 1 USD", fontsize=14, fontweight="bold")
    d = diario.set_index("fecha")
    ejes = ax.flat
    for i, label in enumerate(monedas):
        a = ejes[i]
        a.plot(d.index, d[label], color="#2E5496")
        a.set_title(label, fontsize=10)
        a.grid(True, alpha=0.3)
        a.tick_params(labelsize=8)
    for j in range(len(monedas), len(ejes)):     # apago los ejes que me sobran
        ejes[j].axis("off")
    fig.text(0.99, 0.01,
             f"Fuente: Frankfurter (tipos de referencia del Banco Central Europeo).  "
             f"Generado {dt.date.today():%Y-%m-%d}.",
             ha="right", va="bottom", fontsize=7.5, color="#808080")
    fig.tight_layout(rect=[0, 0.02, 1, 0.97])
    fig.savefig(ruta_png, dpi=130)
    plt.close(fig)


# ------------------------------------------------------------------
# 3b) Exportar a JSON (esto es lo que lee el panel / GitHub Pages)
# ------------------------------------------------------------------
def escribir_json(diario, macro, ruta, acereras=None, info_acereras=None,
                  construccion=None, chatarra=None, productores=None):
    """Tira todo a un JSON compacto que es lo que consume el index.html."""
    def limpia(serie, dec=6):
        out = []
        for v in serie:
            out.append(None if pd.isna(v) else round(float(v), dec))
        return out

    series = {col: limpia(diario[col]) for col in diario.columns if col != "fecha"}

    resumen = []
    for nombre, fmt in _series_resumen():
        if nombre not in diario.columns:
            continue
        sub = diario[["fecha", nombre]].dropna()
        if sub.empty:
            continue
        ult = sub.iloc[-1]
        resumen.append({"serie": nombre, "valor": round(float(ult[nombre]), 6),
                        "fecha": ult["fecha"].strftime("%Y-%m-%d"), "fmt": fmt})

    # indicadores macro (mensuales/trimestrales). 'escala' me lleva el valor a %.
    macro_def = [
        ("Inflación YoY (CPI)",        "Inflacion YoY", 100),
        ("Desempleo",                  "Desempleo",       1),
        ("PIB real (crec. anualizado)", "PIB crec.",      1),
    ]
    macro_out = []
    for nombre, col, escala in macro_def:
        if col not in macro.columns:
            continue
        sub = macro[["fecha", col]].dropna()
        if sub.empty:
            continue
        macro_out.append({
            "nombre": nombre, "escala": escala,
            "fecha": sub["fecha"].dt.strftime("%Y-%m-%d").tolist(),
            "valores": limpia(sub[col], 5),
        })

    # indicadores de construcción (niveles mensuales). Cada uno trae su unidad y
    # formato para que el panel lo muestre tal cual (estos NO son porcentajes).
    construccion_def = [
        # (nombre, columna, unidad, decimales, prefijo, sufijo)
        ("Housing Starts",        "Housing Starts",
         "miles de unidades · SAAR", 0, "",  ""),
        ("Construction Spending", "Construction Spending",
         "mil millones USD · SAAR",  0, "$", ""),
    ]
    construccion_out = []
    if construccion is not None and not construccion.empty:
        for nombre, col, unidad, dec, prefijo, sufijo in construccion_def:
            if col not in construccion.columns:
                continue
            sub = construccion[["fecha", col]].dropna()
            if sub.empty:
                continue
            construccion_out.append({
                "nombre": nombre, "unidad": unidad, "dec": dec,
                "prefijo": prefijo, "sufijo": sufijo,
                "fecha": sub["fecha"].dt.strftime("%Y-%m-%d").tolist(),
                "valores": limpia(sub[col], 1),
            })

    # chatarra ferrosa (índices de precio mensuales del PPI). El nivel es un
    # índice base 1982=100 (no $/ton): por eso la tarjeta resalta el % mensual,
    # que es la lectura útil del mercado.
    chatarra_out = []
    if chatarra is not None and not chatarra.empty:
        for col in chatarra.columns:
            if col == "fecha":
                continue
            sub = chatarra[["fecha", col]].dropna()
            if sub.empty:
                continue
            chatarra_out.append({
                "nombre": col, "unidad": "índice 1982=100", "dec": 1,
                "prefijo": "", "sufijo": "",
                "fecha": sub["fecha"].dt.strftime("%Y-%m-%d").tolist(),
                "valores": limpia(sub[col], 1),
            })

    ahora_utc = dt.datetime.now(dt.timezone.utc)
    obj = {
        "generado": ahora_utc.strftime("%Y-%m-%d %H:%M UTC"),
        "generado_iso": ahora_utc.strftime("%Y-%m-%dT%H:%M:%SZ"),
        "fuentes": "Henry Hub: Yahoo (futuro front-month NG=F), respaldo EIA · "
                   "S&P 500, WTI, Brent, inflación, desempleo, PIB, "
                   "Housing Starts, Construction Spending, Chatarra (PPI): FRED · "
                   "Divisas: Frankfurter (BCE) · Acciones: Twelve Data y Yahoo",
        "fechas": diario["fecha"].dt.strftime("%Y-%m-%d").tolist(),
        "series": series,
        "monedas": [label for (_iso, (label, _f)) in MONEDAS.items()],
        "macro": macro_out,
        "construccion": construccion_out,
        "chatarra": chatarra_out,
        "resumen": resumen,
    }

    if acereras is not None and not acereras.empty:
        ia = info_acereras or {}
        cg = ia.get("corr_global")
        # La columna del índice real puede llamarse "NYSE Arca Steel Index" (si
        # ^STEEL sí respondió) o "NYSE Arca Steel Index (SLX ETF)" (respaldo).
        # La detecto por descarte, ya que es la única columna extra aparte de
        # fecha/mundial/EE. UU.
        cols_base = {"fecha", "Índice mundial", "Índice EE. UU."}
        col_indice_real = next((c for c in acereras.columns if c not in cols_base), None)
        obj["acereras"] = {
            "fecha": acereras["fecha"].dt.strftime("%Y-%m-%d").tolist(),
            "mundial": limpia(acereras["Índice mundial"], 2),
            "eeuu": limpia(acereras["Índice EE. UU."], 2),
            "indice_real": (limpia(acereras[col_indice_real], 2)
                            if col_indice_real else None),
            "indice_real_nombre": (col_indice_real or "NYSE Arca Steel Index"),
            "corr_fecha": ia.get("corr_fecha", []),
            "corr": ia.get("corr_series", []),
            "corr_global": (None if cg is None else round(cg, 3)),
            "ok_mundial": ia.get("ok_mundial", []),
            "ok_eeuu": ia.get("ok_eeuu", []),
            "fallaron": ia.get("fallaron", []),
            "precios": ia.get("precios"),
        }

    # Top 10 productores de acero (worldsteel). Es un ranking chico, lo meto tal cual.
    if productores:
        obj["productores"] = productores

    with open(ruta, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False)


# ------------------------------------------------------------------
# 4) El programa principal
# ------------------------------------------------------------------
def main():
    if not EIA_API_KEY or not FRED_API_KEY:
        log("ERROR: faltan llaves EIA_API_KEY y/o FRED_API_KEY "
            "(.env en tu PC, o Secrets en GitHub Actions).")
        sys.exit(1)

    log("===== Inicio =====")
    try:
        diario = construir_diario()
        log("Descargando indicadores macro (FRED)...")
        macro = obtener_macro()

        log("Descargando indicadores de construcción (FRED)...")
        try:
            construccion = obtener_construccion()
        except Exception as e:
            log(f"  Aviso: la sección de construcción falló y se omite: {e}")
            construccion = None

        log("Descargando chatarra (FRED / PPI)...")
        try:
            chatarra = obtener_chatarra()
        except Exception as e:
            log(f"  Aviso: la sección de chatarra falló y se omite: {e}")
            chatarra = None

        log("Descargando acereras (Twelve Data + Yahoo)...")
        try:
            acereras, info_ac = construir_acereras()
        except Exception as e:
            log(f"  Aviso: la sección de acereras falló y se omite: {e}")
            acereras, info_ac = None, None

        log("Descargando Top 10 productores de acero (worldsteel)...")
        try:
            productores = obtener_productores()
        except Exception as e:
            log(f"  Aviso: la sección de productores falló y se omite: {e}")
            productores = None

        xlsx_fecha = CARPETA / f"mercados_{HOY}.xlsx"
        xlsx_recie = CARPETA / "mercados_reciente.xlsx"
        png_fecha = CARPETA / f"graficos_{HOY}.png"
        png_recie = CARPETA / "graficos_reciente.png"
        png_fx_fecha = CARPETA / f"graficos_divisas_{HOY}.png"
        png_fx_recie = CARPETA / "graficos_divisas_reciente.png"

        escribir_excel(diario, macro, xlsx_fecha, acereras=acereras)
        escribir_excel(diario, macro, xlsx_recie, acereras=acereras)
        crear_graficos(diario, macro, png_fecha)
        crear_graficos(diario, macro, png_recie)
        crear_graficos_fx(diario, png_fx_fecha)
        crear_graficos_fx(diario, png_fx_recie)

        # los datos para el panel (GitHub Pages)
        escribir_json(diario, macro, CARPETA / "datos.json",
                      acereras=acereras, info_acereras=info_ac,
                      construccion=construccion, chatarra=chatarra,
                      productores=productores)

        ult_hh = diario.dropna(subset=[COL_HH]).iloc[-1]
        log(f"Henry Hub más reciente: ${ult_hh[COL_HH]:.2f}/MMBtu ({ult_hh['fecha'].date()})")
        if "MXN por USD" in diario.columns:
            ult_mxn = diario.dropna(subset=["MXN por USD"]).iloc[-1]
            log(f"Peso más reciente: {ult_mxn['MXN por USD']:.4f} MXN/USD ({ult_mxn['fecha'].date()})")
        inf = macro.dropna(subset=['Inflacion YoY']).iloc[-1]
        log(f"Inflación YoY más reciente: {inf['Inflacion YoY']*100:.1f}% ({inf['fecha'].date()})")
        log(f"Excel guardado en: {xlsx_recie}")
        log("===== Fin OK =====")
    except Exception as e:
        log(f"ERROR: {e}")
        raise


if __name__ == "__main__":
    main()
